import re
import yaml
import datetime as _dt
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color

# -----------------------------
# YAML-safe conversion
# -----------------------------
def _yaml_safe(v):
    if v is None:
        return None
    if isinstance(v, (bool, int, float, str)):
        return v
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return v.isoformat()
    if isinstance(v, bytes):
        return v.hex()
    return str(v)

def _normalize_argb(color_value):
    if not color_value or not isinstance(color_value, str):
        return None
    s = color_value.strip()
    if re.fullmatch(r"[0-9A-Fa-f]{8}", s):
        return s.upper()
    if re.fullmatch(r"[0-9A-Fa-f]{6}", s):
        return ("FF" + s).upper()
    return None

def _safe_int(x):
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, int):
        return x
    if isinstance(x, float):
        return int(x) if x.is_integer() else None
    if isinstance(x, str):
        s = x.strip()
        if re.fullmatch(r"-?\d+", s):
            try:
                return int(s)
            except Exception:
                return None
        return None
    return None

def _safe_float(x):
    if x is None:
        return None
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    if isinstance(x, str):
        s = x.strip()
        try:
            return float(s)
        except Exception:
            return None
    return None

# -----------------------------
# Color serialize/deserialize (keep theme/indexed/tint)
# -----------------------------
def _color_to_dict(c):
    if c is None:
        return None
    return {
        "type": _yaml_safe(getattr(c, "type", None)),
        "rgb": _yaml_safe(getattr(c, "rgb", None)),
        "theme": _yaml_safe(getattr(c, "theme", None)),
        "indexed": _yaml_safe(getattr(c, "indexed", None)),
        "tint": _yaml_safe(getattr(c, "tint", None)),
        "auto": _yaml_safe(getattr(c, "auto", None)),
    }

def _dict_to_color(d):
    if not d or not isinstance(d, dict):
        return None

    theme = _safe_int(d.get("theme"))
    indexed = _safe_int(d.get("indexed"))
    tint = _safe_float(d.get("tint"))
    auto = d.get("auto")
    auto_bool = bool(auto) if isinstance(auto, (bool, int)) else None

    rgb = _normalize_argb(d.get("rgb"))

    if theme is not None:
        c = Color(theme=theme)
        if tint is not None:
            c.tint = tint
        if auto_bool is not None:
            c.auto = auto_bool
        return c

    if indexed is not None:
        c = Color(indexed=indexed)
        if tint is not None:
            c.tint = tint
        if auto_bool is not None:
            c.auto = auto_bool
        return c

    if rgb:
        c = Color(rgb=rgb)
        if tint is not None:
            c.tint = tint
        if auto_bool is not None:
            c.auto = auto_bool
        return c

    return None

# -----------------------------
# Helpers: style serialize/deserialize
# -----------------------------
def _font_to_dict(f: Font):
    if f is None:
        return None
    return {
        "name": _yaml_safe(f.name),
        "sz": _yaml_safe(f.sz),
        "b": _yaml_safe(bool(f.b) if f.b is not None else None),
        "i": _yaml_safe(bool(f.i) if f.i is not None else None),
        "u": _yaml_safe(f.u),
        "color": _color_to_dict(getattr(f, "color", None)),
    }

def _fill_to_dict(fill: PatternFill):
    if fill is None:
        return None
    return {
        "pattern": _yaml_safe(fill.patternType),
        "fg": _color_to_dict(getattr(fill, "fgColor", None)),
        "bg": _color_to_dict(getattr(fill, "bgColor", None)),
    }

def _alignment_to_dict(a: Alignment):
    if a is None:
        return None
    return {
        "h": _yaml_safe(a.horizontal),
        "v": _yaml_safe(a.vertical),
        "wrap": _yaml_safe(bool(a.wrap_text) if a.wrap_text is not None else None),
        "rot": _yaml_safe(a.textRotation),
        "shrink": _yaml_safe(bool(a.shrink_to_fit) if a.shrink_to_fit is not None else None),
    }

def _side_to_dict(s: Side):
    if s is None:
        return None
    return {
        "style": _yaml_safe(s.style),
        "color": _color_to_dict(getattr(s, "color", None)),
    }

def _border_to_dict(b: Border):
    if b is None:
        return None
    return {
        "l": _side_to_dict(b.left),
        "r": _side_to_dict(b.right),
        "t": _side_to_dict(b.top),
        "b": _side_to_dict(b.bottom),
    }

def _style_to_dict(cell):
    return {
        "font": _font_to_dict(cell.font),
        "fill": _fill_to_dict(cell.fill),
        "aln": _alignment_to_dict(cell.alignment),
        "border": _border_to_dict(cell.border),
    }

def _dict_to_font(d):
    if not d:
        return None
    color_obj = _dict_to_color(d.get("color"))
    return Font(
        name=d.get("name"),
        sz=d.get("sz"),
        b=d.get("b"),
        i=d.get("i"),
        u=d.get("u"),
        color=color_obj,
    )

def _dict_to_fill(d):
    if not d:
        return None

    pattern = d.get("pattern")
    fg = _dict_to_color(d.get("fg"))
    bg = _dict_to_color(d.get("bg"))

    # Avoid solid fill with no valid colors (common cause of black blocks)
    if pattern == "solid" and fg is None and bg is None:
        return None

    kwargs = {}
    if pattern:
        kwargs["patternType"] = pattern
    if fg:
        kwargs["fgColor"] = fg
    if bg:
        kwargs["bgColor"] = bg

    return PatternFill(**kwargs) if kwargs else None

def _dict_to_alignment(d):
    if not d:
        return None
    return Alignment(
        horizontal=d.get("h"),
        vertical=d.get("v"),
        wrapText=d.get("wrap"),
        textRotation=d.get("rot"),
        shrinkToFit=d.get("shrink"),
    )

def _dict_to_side(d):
    if not d:
        return Side()

    kwargs = {}
    if d.get("style"):
        kwargs["style"] = d["style"]

    c = _dict_to_color(d.get("color"))
    if c:
        kwargs["color"] = c

    return Side(**kwargs) if kwargs else Side()

def _dict_to_border(d):
    if not d:
        return None
    return Border(
        left=_dict_to_side((d.get("l") or {})),
        right=_dict_to_side((d.get("r") or {})),
        top=_dict_to_side((d.get("t") or {})),
        bottom=_dict_to_side((d.get("b") or {})),
    )

# -----------------------------
# Merged range helpers
# -----------------------------
def _addr_to_rc(addr: str):
    m = re.fullmatch(r"([A-Z]+)(\d+)", addr)
    if not m:
        return None
    col_letters, row_str = m.group(1), m.group(2)
    row = int(row_str)
    col = 0
    for ch in col_letters:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return row, col

def _parse_range_a1(rng: str):
    if ":" not in rng:
        rc = _addr_to_rc(rng)
        if not rc:
            return None
        r, c = rc
        return r, c, r, c
    a, b = rng.split(":")
    rc1 = _addr_to_rc(a)
    rc2 = _addr_to_rc(b)
    if not rc1 or not rc2:
        return None
    r1, c1 = rc1
    r2, c2 = rc2
    return min(r1, r2), min(c1, c2), max(r1, r2), max(c1, c2)

def _build_merged_map(merged_ranges):
    top_left = set()
    non_top_left = set()
    for rng in merged_ranges or []:
        parsed = _parse_range_a1(rng)
        if not parsed:
            continue
        r1, c1, r2, c2 = parsed
        top_left.add((r1, c1))
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if (r, c) != (r1, c1):
                    non_top_left.add((r, c))
    return top_left, non_top_left

# -----------------------------
# Excel -> Markdown (YAML blueprint in MD)
# -----------------------------
def excel_to_markdown(xlsx_path: str, md_path: str):
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)

    bp = {"format": "excel-blueprint-md", "version": 2, "workbook": {"sheets": []}}

    for ws in wb.worksheets:
        freeze = ws.freeze_panes
        if freeze is None:
            freeze_addr = None
        elif isinstance(freeze, str):
            freeze_addr = freeze
        else:
            freeze_addr = getattr(freeze, "coordinate", None)
        if not freeze_addr:
            freeze_addr = None

        sheet = {
            "name": ws.title,
            "view": {"zoom": _yaml_safe(getattr(ws.sheet_view, "zoomScale", None))},
            "freeze": _yaml_safe(freeze_addr),
            "col_widths": {},
            "row_heights": {},
            "merged": [],
            "cells": [],
        }

        for col_letter, dim in ws.column_dimensions.items():
            if dim.width is not None:
                sheet["col_widths"][str(col_letter)] = float(dim.width)

        for row_idx, dim in ws.row_dimensions.items():
            if dim.height is not None:
                sheet["row_heights"][int(row_idx)] = float(dim.height)

        for r in ws.merged_cells.ranges:
            sheet["merged"].append(str(r))

        if ws.max_row and ws.max_column:
            for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
                for c in row:
                    if c.value is None and (c.has_style is False):
                        continue

                    entry = {
                        "addr": c.coordinate,
                        "nf": _yaml_safe(c.number_format),
                        "s": _style_to_dict(c),
                    }

                    if isinstance(c.value, str) and c.value.startswith("="):
                        entry["f"] = c.value
                        entry["v"] = None
                        entry["t"] = "formula"
                    else:
                        entry["v"] = _yaml_safe(c.value)
                        entry["t"] = _yaml_safe(type(c.value).__name__ if c.value is not None else "None")

                    sheet["cells"].append(entry)

        bp["workbook"]["sheets"].append(sheet)

    md = []
    md.append("# Excel Blueprint\n")
    md.append("下面的 YAML 是可逆蓝图：用于从 Excel 导出并可重建回 Excel。\n")
    md.append("```yaml\n")
    md.append(yaml.safe_dump(bp, allow_unicode=True, sort_keys=False))
    md.append("```\n")

    with open(md_path, "w", encoding="utf-8") as f:
        f.write("".join(md))

# -----------------------------
# Markdown -> Excel
# -----------------------------
def _load_blueprint_from_md(md_path: str):
    with open(md_path, "r", encoding="utf-8") as f:
        txt = f.read()
    m = re.search(r"```yaml\s*(.*?)\s*```", txt, re.S)
    if not m:
        raise ValueError("No ```yaml ... ``` block found in markdown.")
    return yaml.safe_load(m.group(1))

def _fix_sheet_view_for_scroll(ws):
    """
    Make rebuilt sheets scroll normally (no frozen panes, no split panes,
    no scrollArea restrictions, sane active cell/top-left).
    """
    # Cancel panes/splits
    ws.freeze_panes = None
    ws.sheet_view.pane = None
    ws.sheet_view.split = None

    # Reset view anchors
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.activeCell = "A1"
    ws.sheet_view.activeCellId = 0
    ws.sheet_view.view = "normal"

    # Clear selections that might pin view
    try:
        ws.sheet_view.selection = []
    except Exception:
        pass

    # Ensure no scrollArea restriction (best-effort; attribute differs by version)
    if hasattr(ws, "scrollArea"):
        try:
            ws.scrollArea = None
        except Exception:
            pass

def markdown_to_excel(md_path: str, xlsx_out_path: str):
    bp = _load_blueprint_from_md(md_path)

    wb = openpyxl.Workbook()
    wb.remove(wb.worksheets[0])

    for s in bp["workbook"]["sheets"]:
        ws = wb.create_sheet(s["name"])

        # Fix: allow scrolling for all sheets + cancel all freezes 
        _fix_sheet_view_for_scroll(ws)

        merged_ranges = s.get("merged") or []
        _, non_top_left_set = _build_merged_map(merged_ranges)

        zoom = (s.get("view") or {}).get("zoom")
        if zoom:
            ws.sheet_view.zoomScale = int(zoom)

        for col_letter, w in (s.get("col_widths") or {}).items():
            ws.column_dimensions[col_letter].width = float(w)

        for r, h in (s.get("row_heights") or {}).items():
            ws.row_dimensions[int(r)].height = float(h)

        for e in (s.get("cells") or []):
            addr = e["addr"]
            c = ws[addr]

            if e.get("t") == "formula" and e.get("f"):
                c.value = e["f"]
            else:
                c.value = e.get("v")

            if e.get("nf") is not None:
                c.number_format = e["nf"]

            rc = _addr_to_rc(addr)
            if rc and rc in non_top_left_set:
                continue

            st = e.get("s") or {}
            fnt = _dict_to_font(st.get("font"))
            if fnt:
                c.font = fnt
            fi = _dict_to_fill(st.get("fill"))
            if fi:
                c.fill = fi
            al = _dict_to_alignment(st.get("aln"))
            if al:
                c.alignment = al
            bo = _dict_to_border(st.get("border"))
            if bo:
                c.border = bo

        for rng in merged_ranges:
            ws.merge_cells(rng)

        # Re-apply view fix after merges (some Excel clients are picky)
        _fix_sheet_view_for_scroll(ws)

    wb.save(xlsx_out_path)

# -----------------------------
# CLI
# -----------------------------
if __name__ == "__main__":
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("--xlsx", help="input xlsx")
    p.add_argument("--md", help="output md")
    p.add_argument("--from-md", help="input md to rebuild")
    p.add_argument("--xlsx-out", help="output xlsx when rebuilding")

    args = p.parse_args()

    if args.xlsx and args.md:
        excel_to_markdown(args.xlsx, args.md)
        print(f"Exported markdown: {args.md}")

    if args.from_md and args.xlsx_out:
        markdown_to_excel(args.from_md, args.xlsx_out)
        print(f"Rebuilt xlsx: {args.xlsx_out}")
